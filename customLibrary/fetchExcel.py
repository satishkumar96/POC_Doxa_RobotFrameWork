import openpyxl


def fetch_login_excel(sheetname, row, col):
    file_path = './customLibrary/getdata.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.get_sheet_by_name(sheetname)
    cell = worksheet.cell(row, col)
    return cell.value


def put_title(title):
    file_path = './customLibrary/getdata.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name("Inputs")

    c1 = sheet['B1']
    c1.value = title

    wb.save(file_path)


def get_title():
    file_path = './customLibrary/getdata.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name("Inputs")

    c1 = sheet['B1']
    return c1.value


def put_pr_number(pr_number):
    file_path = './customLibrary/getdata.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name("Inputs")

    c1 = sheet['B2']
    c1.value = pr_number

    wb.save(file_path)


def get_pr_number():
    file_path = './customLibrary/getdata.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name("Inputs")

    c1 = sheet['B2']
    return c1.value
